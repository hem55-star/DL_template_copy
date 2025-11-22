# -*- coding: utf-8 -*-
"""
Created on Mon Nov 10 10:33:38 2025

@author: hmallikarachchi
"""

import statistics as stat
import math
import numpy as np
import os
import sys
import pandas as pd
import pwlf
import xlwings as xw
import shutil
import time
import glob


import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from matplotlib.patches import Rectangle
from matplotlib.lines import Line2D
from matplotlib.legend import Legend
from matplotlib import gridspec

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import coordinate_to_tuple

DL_type = 'per_Unit'        # per_Unit,per_Location


####################User defined functions###########################
def paste_dataframe(df,start_row,start_col,ws):
    rows = dataframe_to_rows(df, index=False, header=False)
    for r_idx, row in enumerate(rows, start=start_row):
        col_offset = start_col - 1
        for c_idx, value in enumerate(row):
            cell = ws.cell(row=r_idx, column=c_idx + start_col)
            cell.value = value
             
    return

def paste_series(depth_series,value_series,depth_start_cell,value_start_cell,ws,title_cell,title):
            
    #Define starting rows and columns 
    # depth_start_cell = "AP4"  # CPT correlation 1
    # value_start_cell = "AQ4"     # CPT correlation 2

    # Convert start cell to row/col index    
    depth_row, depth_col = coordinate_to_tuple(depth_start_cell)
    value_row, value_col = coordinate_to_tuple(value_start_cell)

    # Write depth values
    for i, val in enumerate(depth_series):
        ws.cell(row=depth_row + i, column=depth_col, value=val)

    # Write value values
    for i, val in enumerate(value_series):
        ws.cell(row=value_row + i, column=value_col, value=val)
        
    #Legend
    ws[title_cell]=title
    
    return

   
#################################

## INPUTS
df = pd.read_excel('All_CPTs.xlsx',sheet_name='All_CPTs')
depth_bins_df = pd.read_excel('Depth_bins.xlsx')
lab_data_path = r"C:\Users\hmallikarachchi\OneDrive - Venterra Limited\Documents\GIR_tool_git\Python-GIR-GDG\Lab_data_per_units"



## OUTPUTS
DL_plot_path = "./design_lines_from_template/"

# Load the Excel template
template_path = "Su_DL_template.xlsx"



if DL_type == 'per_Location':
    loca_id = 'BAL01_GT1_CPTU_02' # for testing
    
    # Subset data for this location
    df_loc = df[df['LOCA_ID'] == loca_id].copy()
    bins_loc = depth_bins_df[depth_bins_df['LOCA_ID'] == loca_id].copy()
    
    
    #----------------------------------paste------------------------------------       
    wb = load_workbook(template_path)       
    template_sheet = wb.active  
    
    #paste_dataframe(cpt_data1,4,42,template_sheet)  
    paste_series(df_loc['SCPT_DPTH'],df_loc['SCPT_SUHE'],'AP4','AQ4',template_sheet,'AO2', 'Su HE') 
    paste_series(df_loc['SCPT_DPTH'],df_loc['SCPT_SULE'],'AT4','AU4',template_sheet,'AS2', 'Su LE') 

    
     # Save the workbook
    save_path = os.path.join(DL_plot_path, f"Su_{loca_id}.xlsx")
    wb.save(save_path)
    
units = ['GU ID2c','GU ID2d','GU ID2b']

if DL_type == 'per_Unit':
#    unit_id = 'GU ID2c'  # for testing
    for unit_id in units:
    
        # Subset of cpt data for this unit
        df_unit = df[df['SCPT_UNIT'] == unit_id].copy()
    #    bins_unit = depth_bins_df[depth_bins_df['SCPT_UNIT'] == unit_id].copy()
        # Search for files containing the unit name
        search_pattern = os.path.join(lab_data_path, f"*{unit_id}*.xlsx")
        matching_files = glob.glob(search_pattern)
        
        if matching_files:
            selected_unit_file = matching_files[0]  # first match
            print("Found file:", selected_unit_file)
        else:
            print("No file found for unit:", unit_id)
            
        TRIT_SU = pd.read_excel(selected_unit_file,sheet_name='TRIT_SU')
        TRET_CAUC = pd.read_excel(selected_unit_file,sheet_name='TRET_CAUC')
        TRET_CAUE = pd.read_excel(selected_unit_file,sheet_name='TRET_CAUE')
        TRET_CK0U = pd.read_excel(selected_unit_file,sheet_name='TRET_CK0U')
        TRET_CU = pd.read_excel(selected_unit_file,sheet_name='TRET_CU')
        TORV = pd.read_excel(selected_unit_file,sheet_name='TORV')
        LPEN = pd.read_excel(selected_unit_file,sheet_name='LPEN')
        
    #----------------------------------paste------------------------------------        
        wb = load_workbook(template_path)       
        template_sheet = wb.active  
        
        #paste_dataframe(cpt_data1,4,42,template_sheet)  
        paste_series(df_unit['SCPT_DPTH'],df_unit['SCPT_SUHE'],'AP4','AQ4',template_sheet,'AO2', 'Su HE') 
        paste_series(df_unit['SCPT_DPTH'],df_unit['SCPT_SULE'],'AT4','AU4',template_sheet,'AS2', 'Su LE') 
        paste_series(TRIT_SU['SPEC_DPTH'],TRIT_SU['TRIT_CU'],'CD4','CE4',template_sheet,'CC2', 'UU') 
        paste_series(TRET_CAUC['SPEC_DPTH'],TRET_CAUC['TRET_CU'],'CH4','CI4',template_sheet,'CG2', 'CAUC') 
        paste_series(TRET_CAUE['SPEC_DPTH'],TRET_CAUE['TRET_CU'],'CL4','CM4',template_sheet,'CK2', 'CAUE') 
        paste_series(TRET_CK0U['SPEC_DPTH'],TRET_CK0U['TRET_CU'],'CP4','CQ4',template_sheet,'CO2', 'CK0U') 
        paste_series(TRET_CU['SPEC_DPTH'],TRET_CU['TRET_CU'],'CT4','CU4',template_sheet,'CS2', 'CU') 
        if not TORV.empty:paste_series(TORV['SPEC_DPTH'],TORV['TORV_PUSS'],'CX4','CY4',template_sheet,'CW2', 'Torvane') 
        paste_series(LPEN['SPEC_DPTH'],LPEN['LPEN_PPEN'],'CT4','CU4',template_sheet,'CS2', 'Pocket Penetrometer') 
        
        # Save the workbook
        save_path = os.path.join(DL_plot_path, f"Su_{unit_id}.xlsx")
        wb.save(save_path)
