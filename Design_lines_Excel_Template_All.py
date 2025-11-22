# -*- coding: utf-8 -*-
"""
Created on Mon Nov 10 10:33:38 2025

@author: hmallikarachchi
"""

import statistics as stat
import math
import numpy as np
import os
import sys
import pandas as pd
import pwlf
import xlwings as xw
import shutil
import time
import glob


import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from matplotlib.patches import Rectangle
from matplotlib.lines import Line2D
from matplotlib.legend import Legend
from matplotlib import gridspec

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import coordinate_to_tuple


####################User defined functions###########################
def paste_dataframe(df,start_row,start_col,ws):
    rows = dataframe_to_rows(df, index=False, header=False)
    for r_idx, row in enumerate(rows, start=start_row):
        col_offset = start_col - 1
        for c_idx, value in enumerate(row):
            cell = ws.cell(row=r_idx, column=c_idx + start_col)
            cell.value = value
             
    return

def paste_series(depth_series,value_series,depth_start_cell,value_start_cell,ws,title_cell,title):
            
    #Define starting rows and columns 
    # depth_start_cell = "AP4"  # CPT correlation 1
    # value_start_cell = "AQ4"     # CPT correlation 2

    # Convert start cell to row/col index    
    depth_row, depth_col = coordinate_to_tuple(depth_start_cell)
    value_row, value_col = coordinate_to_tuple(value_start_cell)

    # Write depth values
    for i, val in enumerate(depth_series):
        ws.cell(row=depth_row + i, column=depth_col, value=val)

    # Write value values
    for i, val in enumerate(value_series):
        ws.cell(row=value_row + i, column=value_col, value=val)
        
    #Legend
    ws[title_cell]=title
    
    return

   
#################################

## INPUTS
xls = pd.ExcelFile('./DL_input_file.xlsx')
main_sheet = pd.read_excel(xls, 'Main', header=None)
to_plot_cpt = pd.read_excel(xls, 'To_Plot_CPT').dropna()
to_plot_lab = pd.read_excel(xls, 'To_Plot_Lab').dropna()
depth_bins_df = pd.read_excel(xls, 'BH_information')

chart_type = main_sheet.iloc[0,1]
parameter = main_sheet.iloc[1,1]
CPT_parameter_file = main_sheet.iloc[2,1]
CPT_sheet_name = main_sheet.iloc[3,1]
units = main_sheet.iloc[4,1]
calc_methods = main_sheet.iloc[5,1]
lab_data_path = main_sheet.iloc[6,1]
DL_template = main_sheet.iloc[7,1]
DL_plot_path = main_sheet.iloc[8,1]
unit_list = main_sheet.iloc[9,1]
location_list = main_sheet.iloc[10,1]
depth_bin_file = main_sheet.iloc[11,1]

# Read cpt output GIR tool
df = pd.read_excel(CPT_parameter_file,sheet_name='All_CPTs')
# if os.path.exists(depth_bin_file):
#     depth_bins_df = pd.read_excel(depth_bin_file)
# else:
#     depth_bins_df = pd.DataFrame()




if chart_type == 'Per_Location':
    #loca_id = 'BAL01_GT1_CPTU_06' # for testing
    loc_list = [x.strip() for x in location_list.split(",")]
    for loca_id in loc_list:
        print(loca_id)
    
        # Subset data for this location
        df_loc = df[df['LOCA_ID'] == loca_id].copy()
        if not depth_bins_df.empty: bins_loc = depth_bins_df[depth_bins_df['LOCA_ID'] == loca_id].copy()    
        
        #----------------------------------paste------------------------------------       
        wb = load_workbook(DL_template)       
        template_sheet = wb.active  
        
        for row in to_plot_cpt.itertuples():
            print(row.Label, row.X_axis_header, row.Y_axis_header, row.Label_cell,row.X_axis_cell,row.Y_axis_cell)   
            paste_series(df_loc[row.Y_axis_header],df_loc[row.X_axis_header],row.Y_axis_cell,row.X_axis_cell,template_sheet,row.Label_cell, row.Label)  
    
        if not depth_bins_df.empty: paste_series(bins_loc['soil_type'],bins_loc['From'],'C85','D85',template_sheet,'C83', 'Stratum') 
         # Save the workbook
        save_path = os.path.join(DL_plot_path, f"{parameter}_{loca_id}.xlsx")
        wb.save(save_path)
    
#units = ['GU ID2c','GU ID2d','GU ID2b']


if chart_type == 'Per_Unit':
#    unit_id = 'GU ID2c'  # for testing
    unit_list = [x.strip() for x in unit_list.split(",")]
    for unit_id in unit_list:
        print(unit_id)
        # Subset of cpt data for this unit
        df_unit = df[df['SCPT_UNIT'] == unit_id].copy()
    
        # Search for lab data files containing the unit name
        search_pattern = os.path.join(lab_data_path, f"*{unit_id}*.xlsx")
        matching_files = glob.glob(search_pattern)
        
        if matching_files:
            selected_unit_file = matching_files[0]  # first match
            print("Found file:", selected_unit_file)
        else:
            print("No file found for unit:", unit_id)
            
        
    #----------------------------------paste------------------------------------        
        wb = load_workbook(DL_template)       
        template_sheet = wb.active  
        #template_sheet.title = f"Su_{unit_id}"          # renaming tab mess outputs-check
        
        for row in to_plot_cpt.itertuples():
            print(row.Label, row.X_axis_header, row.Y_axis_header, row.Label_cell,row.X_axis_cell,row.Y_axis_cell)
            paste_series(df_unit[row.Y_axis_header],df_unit[row.X_axis_header],row.Y_axis_cell,row.X_axis_cell,template_sheet,row.Label_cell, row.Label) 
        
        
        for row in to_plot_lab.itertuples():
            print(row.Lab_sheet_name,row.Label, row.X_axis_header, row.Y_axis_header, row.Label_cell,row.X_axis_cell,row.Y_axis_cell)
            print(f"'{row.Lab_sheet_name}'")
            LAB_DATA = pd.read_excel(selected_unit_file,sheet_name=row.Lab_sheet_name)
            if not LAB_DATA.empty:paste_series(LAB_DATA[row.Y_axis_header],LAB_DATA[row.X_axis_header],row.Y_axis_cell,row.X_axis_cell,template_sheet,row.Label_cell, row.Label)
        
        # Save the workbook
       
        save_path = os.path.join(DL_plot_path, f"{parameter}_{unit_id}.xlsx")
        wb.save(save_path)
